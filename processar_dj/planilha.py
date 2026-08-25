"""Geração da planilha de Resgate (modelo) a partir da planilha diária de resgates."""

from __future__ import annotations

from copy import copy
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import logging

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger("jupiter.processarDJ")

ABA_MODELO = "RESGATES EM MODELO"
ABA_DIARIA = "RESGATES A FAVOR DO GOVER"

LINHA_CABECALHO_MODELO = 5
PRIMEIRA_LINHA_DADOS = 6

# coluna do modelo -> coluna da planilha diária (mesma linha)
MAPA_COLUNAS = {
    "A": "A",  # PROCESSO           <- NUMERO_DO_PROCESSO
    "B": "D",  # ORGÃO              <- ORGAO
    "C": "H",  # RECLAMADO          <- NOME_RECLAMADO
    "D": "I",  # CNPJ               <- CPF_CNPJ_RECLAMADO
    "E": "J",  # CONTA JUDICIAL     <- CONTA_JUDICIAL
    "F": "N",  # SALDO CAPITAL      <- VALOR_SALDO_CAPITAL
    "G": "O",  # CORREÇÃO           <- VALOR_CORRECAO_MONETARIA
    "H": "P",  # JUROS              <- VALOR_JUROS
    "I": "Q",  # SALDO CORRIGIDO    <- VALOR_SALDO_CORRIGIDO
}
COLUNA_LC151 = "J"
COLUNA_CONTA_JUDICIAL_MODELO = "E"
CELULA_DATA_CABECALHO = "A4"
PLACEHOLDER_DATA_CABECALHO = "xx/xx/xxxx"

PREENCHIMENTO_AMARELO = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")


def _lc151_bruto(saldo_corrigido: float | None) -> Decimal:
    """LC 151 '30%' de uma linha, sem arredondar."""
    return Decimal(str(saldo_corrigido or 0)) * Decimal("0.3")


def _linhas_dados_diario(ws: Worksheet):
    """Itera as linhas de dados da planilha diária, pulando linhas sem número de processo."""
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        yield row


def _copiar_estilo_linha(ws: Worksheet, linha_origem: int, linha_destino: int) -> None:
    """Copia fonte, borda, preenchimento, formato numérico e alinhamento de uma linha para outra."""
    for col in range(1, ws.max_column + 1):
        origem = ws.cell(row=linha_origem, column=col)
        destino = ws.cell(row=linha_destino, column=col)
        destino.font = copy(origem.font)
        destino.border = copy(origem.border)
        destino.fill = copy(origem.fill)
        destino.number_format = origem.number_format
        destino.alignment = copy(origem.alignment)


def gerar_planilha_resgate(
    caminho_template: Path,
    caminho_diario: Path,
    contas_resgatadas: list[str],
    caminho_saida: Path,
    data_pagamento: str | None = None,
) -> float:
    """Preenche o template de Resgate com as linhas da planilha diária,
    pinta de amarelo as linhas cuja CONTA JUDICIAL está em ``contas_resgatadas``
    e preenche a data do cabeçalho com ``data_pagamento``, se informada.

    Retorna o valor total de LC 151 '30%'.
    """
    wb_diario = openpyxl.load_workbook(caminho_diario, data_only=True)
    ws_diario = wb_diario[ABA_DIARIA]

    wb_modelo = openpyxl.load_workbook(caminho_template)
    ws_modelo = wb_modelo[ABA_MODELO]

    if data_pagamento:
        celula_data = ws_modelo[CELULA_DATA_CABECALHO]
        if celula_data.value and PLACEHOLDER_DATA_CABECALHO in str(celula_data.value):
            celula_data.value = str(celula_data.value).replace(PLACEHOLDER_DATA_CABECALHO, data_pagamento)

    linhas_diarias = list(_linhas_dados_diario(ws_diario))
    if not linhas_diarias:
        raise ValueError(f"Nenhuma linha de dados encontrada em {caminho_diario.name}.")
    qtd_linhas = len(linhas_diarias)

    linha_totais_original = ws_modelo.max_row
    qtd_linhas_exemplo = linha_totais_original - PRIMEIRA_LINHA_DADOS

    try:
        ws_modelo.unmerge_cells(f"A{linha_totais_original}:B{linha_totais_original}")
    except KeyError:
        pass

    if qtd_linhas > qtd_linhas_exemplo:
        ws_modelo.insert_rows(PRIMEIRA_LINHA_DADOS + qtd_linhas_exemplo, amount=qtd_linhas - qtd_linhas_exemplo)
    elif qtd_linhas < qtd_linhas_exemplo:
        ws_modelo.delete_rows(PRIMEIRA_LINHA_DADOS + qtd_linhas, amount=qtd_linhas_exemplo - qtd_linhas)

    ultima_linha_dados = PRIMEIRA_LINHA_DADOS + qtd_linhas - 1
    linha_totais = ultima_linha_dados + 1
    ws_modelo.merge_cells(f"A{linha_totais}:B{linha_totais}")

    contas_alvo = {c.strip() for c in contas_resgatadas}

    for offset, linha_diaria in enumerate(linhas_diarias):
        linha_destino = PRIMEIRA_LINHA_DADOS + offset
        _copiar_estilo_linha(ws_modelo, PRIMEIRA_LINHA_DADOS, linha_destino)

        for col_modelo, col_diario in MAPA_COLUNAS.items():
            valor = ws_diario[f"{col_diario}{linha_diaria[0].row}"].value
            ws_modelo[f"{col_modelo}{linha_destino}"] = valor

        ws_modelo[f"{COLUNA_LC151}{linha_destino}"] = f"=I{linha_destino}*0.3"

        conta_judicial = ws_modelo[f"{COLUNA_CONTA_JUDICIAL_MODELO}{linha_destino}"].value
        if conta_judicial is not None and str(conta_judicial).strip() in contas_alvo:
            for col in list(MAPA_COLUNAS) + [COLUNA_LC151]:
                ws_modelo[f"{col}{linha_destino}"].fill = PREENCHIMENTO_AMARELO

    for col in ("F", "G", "H", "I"):
        ws_modelo[f"{col}{linha_totais}"] = f"=SUM({col}{PRIMEIRA_LINHA_DADOS}:{col}{ultima_linha_dados})"
    ws_modelo[f"{COLUNA_LC151}{linha_totais}"] = (
        f"=ROUND(SUM({COLUNA_LC151}{PRIMEIRA_LINHA_DADOS}:{COLUNA_LC151}{ultima_linha_dados}),2)"
    )

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb_modelo.save(caminho_saida)

    soma_bruta = sum(
        (_lc151_bruto(ws_modelo[f"I{PRIMEIRA_LINHA_DADOS + i}"].value) for i in range(qtd_linhas)),
        start=Decimal("0"),
    )
    valor_gr = float(soma_bruta.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    log.info(f"Planilha de Resgate gerada: {caminho_saida.name} ({qtd_linhas} linha(s), GR = R$ {valor_gr:,.2f}).")
    return valor_gr
